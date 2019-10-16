from PySimpleGUI import PopupError,FileBrowse,FilesBrowse,Text,Input,Multiline,Button,Column,Radio,Window,TabGroup,Tab
import multiprocessing
from sys import platform
from openpyxl import load_workbook
from os.path import splitext
import moralstrength as MS
from lexicon_use import form_text_vector
import estimators as estimators
from numpy import hstack

model_to_use = ''

trans_list = dict()
lr_lst = dict() 

#iterate over radio buttons and find which model to use
def getModel(values):
	global model_to_use
#	pdb.set_trace()
	for model in estimators.models:
		if values[model]:
			model_to_use = model

def processTextWithMoral(text,moral):
	transformers = trans_list[moral]
	lr = lr_lst[moral]
	X = []
	for transformer in transformers.keys():
		if transformer == 'unigram':
			X_tmp = transformers[transformer].transform([' '.join(t) for t in text])
			X_tmp = X_tmp.toarray()
			X.append(X_tmp)
		elif transformer == 'simon':
			X_tmp = transformers[transformer].transform(text)
			X.append(X_tmp)
		else: 
			X_tmp = [form_text_vector(t, model=transformer) for t in text]    
			X.append(X_tmp)
	X = hstack(X)
	return lr.predict_proba(X)[:, 1][0]


def processExcelFile(inputfilename,outputfilename):
	wb = load_workbook(filename = inputfilename)
	sheet = wb[wb.sheetnames[0]]
	last_col = sheet.max_column + 1
	#first row is the header
	colidx = last_col
	for moral in MS.moral_options_predictions:
		colidx = colidx + 1
		if moral != "non-moral":
			sheet.cell(row=1, column=colidx).value = moral+"_presence"
			colidx = colidx + 1
			sheet.cell(row=1, column=colidx).value = moral+"_word_avg"		
		else:
			#non-moral is special, we don't want to have 'non-moral_presence' in header
			sheet.cell(row=1, column=colidx).value = moral
	#process remaining file
	for rowidx in range(2,sheet.max_row+1): 
		text = sheet.cell(row=rowidx, column=1).value
		text_processed = estimators.pp_pipe.transform([text])
		colidx = last_col
		for moral in MS.moral_options_predictions:
			colidx = colidx + 1
			sheet.cell(row=rowidx, column=colidx).value = processTextWithMoral(text_processed,moral)
			if moral != 'non-moral':
				colidx = colidx + 1
				sheet.cell(row=rowidx, column=colidx+1).value =  MS.string_average_moral(text,moral)

	wb.save(outputfilename)
	return
	
def processTextFile(inputfilename,outputfilename):
	with open(inputfilename, 'r') as f:
		with open(outputfilename, "w") as outf:
			#write header
			outf.write("Input text")
			for moral in MS.moral_options_predictions:
				if moral != "non-moral":
					outf.write("\t"+moral+"_presence")
					outf.write("\t"+moral+"_word_avg")
				else:
					outf.write("\t"+moral)
			outf.write("\n")
			for line in f:
				line = line.strip()
				text_processed = estimators.pp_pipe.transform([line])
				outf.write(line)
				for moral in MS.moral_options_predictions:
					outf.write("\t"+str(processTextWithMoral(text_processed,moral)))
					if moral != 'non-moral':
						outf.write("\t"+str(MS.string_average_moral(line,moral)))
				outf.write("\n")
	return

def processFiles(filelist):
	files = filelist.split(";")
	window.Element('currentfile').Update("")
	#preload all stuff once and then work line-by-line
	#so large files can also be processed without loading them in-memory
	global lr_lst
	global trans_list
	for moral in MS.moral_options_predictions:
		estim, transformers = estimators.select_processes(model_to_use, moral)
		lr, transformers = estimators.load_models(estim, transformers, moral)
		lr_lst[moral] = lr
		trans_list[moral] = transformers
			
	errorstring = ""
	errorfiles = []
	for file in files:
		try:
			isExcel = False
			window.Element('currentfile').Update("Currently processing " +file)
			basefilename, extension = splitext(file)
			if extension.startswith('.xls'):
				isExcel = True
			outputfile = basefilename + "_MoralStrength"+extension
			if not isExcel:
				processTextFile(file,outputfile)
			else:
				processExcelFile(file,outputfile)
		except:
			continue
			errorfiles.append(file)
			errorstring = " (with errors)"
	window.Element('currentfile').Update("All file(s) processed"+errorstring)
	if len(errorfiles)>0:
		PopupError('The following files could not be processed correctly:'+'\n'.join(errorfiles))

def analyzeText(text):
	results = MS.string_moral_values(text)
	maxmoral = ""
	maxvalue = -1
	for moral in results:
		moral = moral
		if results[moral] > maxvalue:
			maxvalue = results[moral]
			maxmoral = moral
		window.Element(moral+"_result").Update("%.3f" % results[moral])
		if moral != 'non-moral':
			trait_avg = MS.string_average_moral(text,moral)
			if trait_avg == -1:
				window.Element(moral+"_avg").Update("No words in lexicon")
			else:
				window.Element(moral+"_avg").Update("%.3f" % trait_avg)
	window.Element(maxmoral+"_result").Update("%.3f (HIGHEST)" % maxvalue)

if platform == "darwin":
	openFileButton = FileBrowse('Select one file', target="files")
	openManyFilesButton = FilesBrowse('Select multiple files', target="files")
else:
	openFileButton = FileBrowse('Select one file', target="files", file_types=(("Text Files", "*.txt"), ("Excel Files", "*.xlsx"), ("ALL Files", "*.*")))
	openManyFilesButton = FilesBrowse('Select multiple files', target="files", file_types=(("Text Files", "*.txt"), ("Excel Files", "*.xlsx"), ("ALL Files", "*.*")))
	


#The first tab is for direct text entry. It has a big text box, a button, and a 2*6 grid for results + labels
output_values = [[Text('Moral Trait', size=(20, 1)), Text('Trait present', size=(36, 1)),Text('Trait avg.', size=(35, 1)) ],
		[Text('Care/Harm', size=(20, 1)), Input('0',key='care_result', size=(35, 1)), Input('0',key='care_avg', size=(35, 1)) ],
		[Text('Fairness/Cheating', size=(20, 1)), Input('0',key='fairness_result', size=(35, 1)), Input('0',key='fairness_avg', size=(35, 1)) ],
		[Text('Loyalty/Betrayal', size=(20, 1)),	 Input('0',key='loyalty_result', size=(35, 1)), Input('0',key='loyalty_avg', size=(35, 1)) ],
		[Text('Authority/Subversion', size=(20, 1)), Input('0',key='authority_result', size=(35, 1)), Input('0',key='authority_avg', size=(35, 1)) ],
		[Text('Purity/Degradation', size=(20, 1)), Input('0',key='purity_result', size=(35, 1)), Input('0',key='purity_avg', size=(35, 1)) ],
		[Text('Non-moral', size=(20, 1)), Input('0',key='non-moral_result',size=(35,1))]]
tab1_layout =  [[Text('Enter text to annotate:')],	  
	[Multiline('',size=(88, 20), key='inputtext')],		
	[Button('Analyze text')],
	[Text(
    '''The output in the first column is the estimated probability of the text being relevant to either a vice or virtue of the corresponding moral trait.
Since the system is trained on tweets, try not to analyze a long text!
The second column is the average rating for the words, based on human ratings of the Moral Foundation Dictionary words.
The range goes from 1 to 9:
1: words closely associated to harm, cheating, betrayal, subversion, degradation
9: words closely associated to care, fairness, loyalty, authority, sanctity'''
    )],
	[Column(output_values)]]	   
	
#the second tab is for analyzing one or more files	
tab2_layout = [[Text(
'''Select one or more files to analyze. If you select a text file, the file results are calculated *per line*.
The output file will contain the input text lines, the 6 predictions and 5 averages, separated by tabs.
	
You can also select an Excel file; the results are calculated on the text in the first column, row by row.
Rrow A is considered a header and *is ignored*.
The results are put in the first available column, and a new file is saved.

The output will always be saved in "[old_filename]_MoralStrength", and the file will be overwritten silently!

As in the other tab, the output is both the estimated probability of the text being relevant to either
a vice or virtue of the corresponding moral trait, and the average of word scores for each trait (in a range from 1 to 9). 
If no word is in the lexicon, the average is set to -1.''')],		
	[openFileButton, openManyFilesButton] ,
	[Text('Selected file(s):')],
	[Text('', key='files', size=(88, 2))],
	[Button('Analyze file(s)')],
	[Text('', key='currentfile', size=(88, 1))],
	]

## the third tab is for configuring the options (i.e., which model to use)
tab3_layout =  [[Text('Select which model should be used for predicting the moral texts:')],
	[Radio('simon', key='simon', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('unigram', key='unigram', group_id="ModelChoiceRadios")],
	[Radio('count', key='count', group_id="ModelChoiceRadios")],
	[Radio('freq', key='freq', group_id="ModelChoiceRadios")],
	[Radio('simon+count', key='simon+count', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('simon+freq', key='simon+freq', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('simon+count+freq', key='simon+count+freq', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('unigram+count', key='unigram+count', group_id="ModelChoiceRadios")],
	[Radio('unigram+freq', key='unigram+freq', group_id="ModelChoiceRadios", default=True)],
	[Radio('unigram+count+freq', key='unigram+count+freq', group_id="ModelChoiceRadios")],
	[Radio('simon+unigram+count', key='simon+unigram+count', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('simon+unigram+freq', key='simon+unigram+freq', group_id="ModelChoiceRadios", disabled=True)],
	[Radio('simon+unigram+count+freq', key='simon+unigram+count+freq', group_id="ModelChoiceRadios", disabled=True)],
	[Text('''You can read what each model does in our paper:
MoralStrength: Exploiting a Moral Lexicon and Embedding Similarity for Moral Foundations Prediction
Oscar Araque, Lorenzo Gatti, Kyriaki Kalimeri
(currently in review at Expert Systems with Applications, but available at https://arxiv.org/abs/1904.08314)

The "simon" models are not available in this GUI, since they require an embeddings file.''')]]	 


layout = [[TabGroup([
			[Tab('Direct text entry', tab1_layout, tooltip='Here you can input some text directly and get a prediction, it is useful for interactive testing'), 
			 Tab('Work on files', tab2_layout, tooltip='Here you can select one or more files to be processed, which is useful for batch works'),
			 Tab('Model selection', tab3_layout, tooltip='Here you can decide which ML model to use for the prediction')]])
						]
		]





# Create the Window
window = Window('MoralStrength GUI', layout)
# Event Loop to process "events"
while True:				
	event, values = window.Read()
	if event in (None, 'Cancel'):
		break
	if event == 'Analyze text':
#		getModel(values)	
		if "inputtext" in values and not values["inputtext"].isspace():
			analyzeText(values["inputtext"])
		else:
			window.Element('inputtext').Update("Write or paste some text first!")
	if event == "Analyze file(s)":
		getModel(values)	
		#ugh... why is the target= of the open file not working? I don't know
		if 'text' in window.Element('files').TKText.keys() and not window.Element('files').TKText['text'].isspace():
			processFiles(window.Element('files').TKText['text'])
		else:
			window.Element('files').Update("Select one (or more) files first!")
window.Close()
